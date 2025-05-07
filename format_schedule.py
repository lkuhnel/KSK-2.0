import pandas as pd
from datetime import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_calendar_sheet(wb, month_date, schedule_df):
    # Create new sheet with month name
    month_name = month_date.strftime("%B %Y")
    ws = wb.create_sheet(title=month_name)
    
    # Write month and year in row 1
    ws.cell(row=1, column=1, value=month_name)
    
    # Column mappings for days of week (each day gets 2 columns)
    day_columns = {
        'Sunday': ['A', 'B'],
        'Monday': ['C', 'D'],
        'Tuesday': ['E', 'F'],
        'Wednesday': ['G', 'H'],
        'Thursday': ['I', 'J'],
        'Friday': ['K', 'L'],
        'Saturday': ['M', 'N']
    }
    
    # Set column widths
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 5

    # Define border styles
    thin_border = Side(style='thin')
    day_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )

    # Write day headers in row 2
    for day, cols in day_columns.items():
        col_idx = ord(cols[0]) - ord('A') + 1
        for offset in range(2):
            cell = ws.cell(row=2, column=col_idx + offset)
            cell.value = day if offset == 0 else ""
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Colors
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')     # Light gray
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')    # Light green
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Light yellow
    
    # Get first day of the month (0 = Monday, 6 = Sunday)
    first_day = datetime(month_date.year, month_date.month, 1).weekday()
    # Convert to Sunday = 0 format
    first_day = (first_day + 1) % 7
    
    # Get number of days in the month
    _, num_days = calendar.monthrange(month_date.year, month_date.month)
    
    # Calculate calendar weeks
    current_week = []
    calendar_weeks = []
    day_count = 0
    
    # Add empty days before the first of the month
    for _ in range(first_day):
        current_week.append(0)
        day_count += 1
    
    # Add all days of the month
    for day in range(1, num_days + 1):
        current_week.append(day)
        day_count += 1
        if day_count % 7 == 0:
            calendar_weeks.append(current_week)
            current_week = []
    
    # Add empty days to complete the last week if needed
    if current_week:
        while len(current_week) < 7:
            current_week.append(0)
        calendar_weeks.append(current_week)
    
    # Write calendar
    for week_num, week in enumerate(calendar_weeks):
        base_row = 3 + (week_num * 8)  # Each week takes 8 rows
        
        # Add week separator border if not first week
        if week_num > 0:
            for col in range(1, 16):  # A through O
                cell = ws.cell(row=base_row - 1, column=col)
                cell.border = Border(bottom=Side(style='thin'))
        
        # Color the rows for each week
        for col in range(1, 16):  # A through O (including column O)
            # Rows 4-5 (gray)
            ws.cell(row=base_row + 3, column=col).fill = gray_fill
            ws.cell(row=base_row + 4, column=col).fill = gray_fill
            
            # Row 7 (green)
            ws.cell(row=base_row + 6, column=col).fill = green_fill
            
            # Row 8 (yellow)
            ws.cell(row=base_row + 7, column=col).fill = yellow_fill
        
        # Add labels in column O for each week
        ws.cell(row=base_row + 3, column=15, value="On Call")
        ws.cell(row=base_row + 6, column=15, value="Supervising")
        ws.cell(row=base_row + 7, column=15, value="Backup")
        
        for weekday, day in enumerate(week):
            if day != 0:
                # Get the starting column for this day
                cols = list(day_columns.values())[weekday]
                col_idx = ord(cols[0]) - ord('A') + 1
                
                # Get schedule for this day
                current_date = datetime(month_date.year, month_date.month, day)
                date_str = current_date.strftime("%Y-%m-%d")
                day_schedule = schedule_df[schedule_df['Date'] == date_str]
                
                # Write day number (only in first column)
                day_cell = ws.cell(row=base_row, column=col_idx)
                day_cell.value = day
                day_cell.alignment = Alignment(horizontal='center')
                
                if not day_schedule.empty:
                    # Write Call person (row 4, first column only)
                    call_cell = ws.cell(row=base_row + 3, column=col_idx)
                    call_cell.value = day_schedule.iloc[0]['Call']
                    call_cell.alignment = Alignment(horizontal='center')
                    
                    # Write Backup person (row 8, first column only)
                    backup_cell = ws.cell(row=base_row + 7, column=col_idx)
                    backup_cell.value = day_schedule.iloc[0]['Backup']
                    backup_cell.alignment = Alignment(horizontal='center')
                    
                    # Write Intern person (row 5, first column only) if assigned
                    if pd.notna(day_schedule.iloc[0]['Intern']):
                        intern_cell = ws.cell(row=base_row + 4, column=col_idx)
                        intern_cell.value = day_schedule.iloc[0]['Intern']
                        intern_cell.alignment = Alignment(horizontal='center')
                        # Style intern cell differently
                        intern_cell.font = Font(italic=True)
                        intern_cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                
                # Add borders around the entire 2x8 day block
                # Apply all borders in a single pass
                for row_offset in range(8):
                    for col_offset in range(2):
                        current_cell = ws.cell(row=base_row + row_offset, column=col_idx + col_offset)
                        
                        # Determine which borders this cell should have
                        has_left = (col_offset == 0)
                        has_right = (col_offset == 1)
                        has_top = (row_offset == 0)
                        has_bottom = (row_offset == 7)
                        
                        # Create border object with all required sides
                        current_cell.border = Border(
                            left=thin_border if has_left else None,
                            right=thin_border if has_right else None,
                            top=thin_border if has_top else None,
                            bottom=thin_border if has_bottom else None
                        )

                # Remove any week separator borders that might interfere with day blocks
                if week_num > 0:
                    for col_offset in range(2):
                        separator_cell = ws.cell(row=base_row - 1, column=col_idx + col_offset)
                        separator_cell.border = Border()

    return ws

# Read the schedule
df = pd.read_csv('generated_schedule.csv')
df['Date'] = pd.to_datetime(df['Date'])

# Create workbook
wb = Workbook()

# Process each month
for month in pd.date_range(df['Date'].min(), df['Date'].max(), freq='MS'):
    month_df = df[df['Date'].dt.month == month.month]
    create_calendar_sheet(wb, month, month_df)

# Remove the default sheet
wb.remove(wb['Sheet'])

# Save the workbook
wb.save('formatted_schedule.xlsx')
print("Calendar-style schedule has been saved to 'formatted_schedule.xlsx'") 