import pandas as pd
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_calendar_sheet(wb, month_date, schedule_df):
    # Create new sheet with month name
    month_name = month_date.strftime("%B %Y")
    ws = wb.create_sheet(title=month_name)
    
    # Write month and year in row 1
    ws.cell(row=1, column=1, value=month_name)
    
    # Column mappings for days of week (each day gets 2 columns)
    day_columns = {
        'Sunday': ['A', 'B'],
        'Monday': ['C', 'D'],
        'Tuesday': ['E', 'F'],
        'Wednesday': ['G', 'H'],
        'Thursday': ['I', 'J'],
        'Friday': ['K', 'L'],
        'Saturday': ['M', 'N']
    }
    
    # Set column widths
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 5

    # Define border styles
    thin_border = Side(style='thin')
    day_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )

    # Write day headers in row 2
    for day, cols in day_columns.items():
        col_idx = ord(cols[0]) - ord('A') + 1
        for offset in range(2):
            cell = ws.cell(row=2, column=col_idx + offset)
            cell.value = day if offset == 0 else ""
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Colors
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')     # Light gray
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')    # Light green
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Light yellow
    
    # Get first and last day of the month
    first_of_month = datetime(month_date.year, month_date.month, 1)
    _, num_days = calendar.monthrange(month_date.year, month_date.month)
    last_of_month = datetime(month_date.year, month_date.month, num_days)

    # Find the first Sunday before or on the 1st
    first_sunday = first_of_month
    while first_sunday.weekday() != 6:  # 6 = Sunday
        first_sunday -= timedelta(days=1)
    # Find the last Saturday after or on the last day
    last_saturday = last_of_month
    while last_saturday.weekday() != 5:  # 5 = Saturday
        last_saturday += timedelta(days=1)

    # Build list of all days to display
    num_days_display = (last_saturday - first_sunday).days + 1
    all_days = [first_sunday + timedelta(days=i) for i in range(num_days_display)]

    # Build weeks
    calendar_weeks = [all_days[i:i+7] for i in range(0, len(all_days), 7)]

    # Write calendar
    for week_num, week in enumerate(calendar_weeks):
        base_row = 3 + (week_num * 8)  # Each week takes 8 rows
        
        # Add week separator border if not first week
        if week_num > 0:
            for col in range(1, 16):  # A through O
                cell = ws.cell(row=base_row - 1, column=col)
                cell.border = Border(bottom=Side(style='thin'))
        
        # Color the rows for each week
        for col in range(1, 16):  # A through O (including column O)
            # Rows 4-5 (gray)
            ws.cell(row=base_row + 3, column=col).fill = gray_fill
            ws.cell(row=base_row + 4, column=col).fill = gray_fill
            # Row 7 (green)
            ws.cell(row=base_row + 6, column=col).fill = green_fill
            # Row 8 (yellow)
            ws.cell(row=base_row + 7, column=col).fill = yellow_fill
        
        # Add labels in column O for each week
        ws.cell(row=base_row + 3, column=15, value="On Call")
        ws.cell(row=base_row + 4, column=15, value="Intern")  # Add intern label
        ws.cell(row=base_row + 6, column=15, value="Supervisor")  # Supervisor label above backup
        ws.cell(row=base_row + 7, column=15, value="Backup")
        
        for weekday, date in enumerate(week):
            cols = list(day_columns.values())[weekday]
            col_idx = ord(cols[0]) - ord('A') + 1
            # Write day number (only in first column)
            day_cell = ws.cell(row=base_row, column=col_idx)
            day_cell.value = date.day
            day_cell.alignment = Alignment(horizontal='center')
            # Italicize if not in current month
            if date.month != month_date.month:
                day_cell.font = Font(italic=True)
            # Get schedule for this day (including overflow days)
            date_str = date.strftime("%Y-%m-%d")
            day_schedule = schedule_df[schedule_df['Date'] == date_str]
            if not day_schedule.empty:
                # Write Call person (row 4, first column only)
                call_cell = ws.cell(row=base_row + 3, column=col_idx)
                call_cell.value = day_schedule.iloc[0].get('Call', '')
                call_cell.alignment = Alignment(horizontal='center')
                # Write Intern person (row 5, first column only)
                intern_cell = ws.cell(row=base_row + 4, column=col_idx)
                intern_cell.value = day_schedule.iloc[0].get('Intern', '')
                intern_cell.alignment = Alignment(horizontal='center')
                # Write Supervisor (row 7, first column only)
                if 'Supervisor' in day_schedule.columns:
                    supervisor_cell = ws.cell(row=base_row + 6, column=col_idx)
                    supervisor_cell.value = day_schedule.iloc[0].get('Supervisor', '')
                    supervisor_cell.alignment = Alignment(horizontal='center')
                # Write Backup person (row 8, first column only)
                backup_cell = ws.cell(row=base_row + 7, column=col_idx)
                backup_cell.value = day_schedule.iloc[0].get('Backup', '')
                backup_cell.alignment = Alignment(horizontal='center')
            # Add borders around the entire 2x8 day block
            for row_offset in range(8):
                for col_offset in range(2):
                    current_cell = ws.cell(row=base_row + row_offset, column=col_idx + col_offset)
                    has_left = (col_offset == 0)
                    has_right = (col_offset == 1)
                    has_top = (row_offset == 0)
                    has_bottom = (row_offset == 7)
                    current_cell.border = Border(
                        left=thin_border if has_left else None,
                        right=thin_border if has_right else None,
                        top=thin_border if has_top else None,
                        bottom=thin_border if has_bottom else None
                    )
            # Remove any week separator borders that might interfere with day blocks
            if week_num > 0:
                for col_offset in range(2):
                    separator_cell = ws.cell(row=base_row - 1, column=col_idx + col_offset)
                    separator_cell.border = Border()
    return ws

def format_schedule(block1_df, block2_df, block3_df):
    # Combine the schedules
    all_df = pd.concat([block1_df, block2_df, block3_df], ignore_index=True)
    all_df['Date'] = pd.to_datetime(all_df['Date'])

    # Create workbook
    wb = Workbook()

    # Get all unique months in the schedule
    months = pd.date_range(all_df['Date'].min(), all_df['Date'].max(), freq='MS')
    
    # Process each month
    for month in months:
        # Get first and last day of the month
        first_of_month = datetime(month.year, month.month, 1)
        _, num_days = calendar.monthrange(month.year, month.month)
        last_of_month = datetime(month.year, month.month, num_days)
        # Find the first Sunday before or on the 1st
        first_sunday = first_of_month
        while first_sunday.weekday() != 6:  # 6 = Sunday
            first_sunday -= timedelta(days=1)
        # Find the last Saturday after or on the last day
        last_saturday = last_of_month
        while last_saturday.weekday() != 5:  # 5 = Saturday
            last_saturday += timedelta(days=1)
        # Get all assignments for the full calendar grid (including overflow days)
        mask = (all_df['Date'] >= first_sunday) & (all_df['Date'] <= last_saturday)
        month_df = all_df[mask]
        create_calendar_sheet(wb, month, month_df)

    # Remove the default sheet
    wb.remove(wb['Sheet'])
    
    return wb

def create_merged_calendar_sheet(wb, prev_month, current_month, schedule_df):
    """Create a calendar sheet that includes the last week of previous month and current month"""
    # Use current month for sheet name
    month_name = current_month.strftime("%B %Y")
    ws = wb.create_sheet(title=month_name)
    
    # Write month and year in row 1
    ws.cell(row=1, column=1, value=month_name)
    
    # Column mappings for days of week (each day gets 2 columns)
    day_columns = {
        'Sunday': ['A', 'B'],
        'Monday': ['C', 'D'],
        'Tuesday': ['E', 'F'],
        'Wednesday': ['G', 'H'],
        'Thursday': ['I', 'J'],
        'Friday': ['K', 'L'],
        'Saturday': ['M', 'N']
    }
    
    # Set column widths
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 5

    # Define border styles
    thin_border = Side(style='thin')
    
    # Write day headers in row 2
    for day, cols in day_columns.items():
        col_idx = ord(cols[0]) - ord('A') + 1
        for offset in range(2):
            cell = ws.cell(row=2, column=col_idx + offset)
            cell.value = day if offset == 0 else ""
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Colors
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Get the earliest date in the schedule
    start_date = schedule_df['Date'].min()
    
    # Calculate the first Sunday of our calendar view
    first_sunday = start_date - pd.Timedelta(days=start_date.weekday() + 1)
    if first_sunday.month != start_date.month:
        first_sunday = start_date - pd.Timedelta(days=start_date.weekday())
    
    # Get number of days to show
    last_date = schedule_df['Date'].max()
    num_days = (last_date - first_sunday).days + 1
    
    # Calculate calendar weeks
    current_week = []
    calendar_weeks = []
    current_date = first_sunday
    
    for _ in range(num_days):
        current_week.append(current_date)
        if len(current_week) == 7:
            calendar_weeks.append(current_week)
            current_week = []
        current_date += pd.Timedelta(days=1)
    
    # Add remaining days if any
    if current_week:
        while len(current_week) < 7:
            current_week.append(current_date)
            current_date += pd.Timedelta(days=1)
        calendar_weeks.append(current_week)
    
    # Write calendar
    for week_num, week in enumerate(calendar_weeks):
        base_row = 3 + (week_num * 8)  # Each week takes 8 rows
        
        # Add week separator border if not first week
        if week_num > 0:
            for col in range(1, 16):  # A through O
                cell = ws.cell(row=base_row - 1, column=col)
                cell.border = Border(bottom=Side(style='thin'))
        
        # Color the rows for each week
        for col in range(1, 16):  # A through O (including column O)
            # Rows 4-5 (gray)
            ws.cell(row=base_row + 3, column=col).fill = gray_fill
            ws.cell(row=base_row + 4, column=col).fill = gray_fill
            
            # Row 7 (green)
            ws.cell(row=base_row + 6, column=col).fill = green_fill
            
            # Row 8 (yellow)
            ws.cell(row=base_row + 7, column=col).fill = yellow_fill
        
        # Add labels in column O for each week
        ws.cell(row=base_row + 3, column=15, value="On Call")
        ws.cell(row=base_row + 4, column=15, value="Intern")
        ws.cell(row=base_row + 6, column=15, value="Supervisor")
        ws.cell(row=base_row + 7, column=15, value="Backup")
        
        for weekday, date in enumerate(week):
            # Get the starting column for this day
            cols = list(day_columns.values())[weekday]
            col_idx = ord(cols[0]) - ord('A') + 1
            
            # Write day number (only in first column)
            day_cell = ws.cell(row=base_row, column=col_idx)
            day_cell.value = date.day
            day_cell.alignment = Alignment(horizontal='center')
            
            # Style differently if from previous month
            if date.month != current_month.month:
                day_cell.font = Font(italic=True, color='808080')
            
            # Get schedule for this day
            date_str = date.strftime("%Y-%m-%d")
            day_schedule = schedule_df[schedule_df['Date'] == date_str]
            
            if not day_schedule.empty:
                # Write Call person (row 4, first column only)
                call_cell = ws.cell(row=base_row + 3, column=col_idx)
                call_cell.value = day_schedule.iloc[0]['Call']
                call_cell.alignment = Alignment(horizontal='center')
                
                # Write Intern person (row 5, first column only)
                intern_cell = ws.cell(row=base_row + 4, column=col_idx)
                intern_cell.value = day_schedule.iloc[0]['Intern']
                intern_cell.alignment = Alignment(horizontal='center')
                
                # Write Supervisor (row 7, first column only)
                if 'Supervisor' in day_schedule.columns:
                    supervisor_cell = ws.cell(row=base_row + 6, column=col_idx)
                    supervisor_cell.value = day_schedule.iloc[0]['Supervisor']
                    supervisor_cell.alignment = Alignment(horizontal='center')
                
                # Write Backup person (row 8, first column only)
                backup_cell = ws.cell(row=base_row + 7, column=col_idx)
                backup_cell.value = day_schedule.iloc[0]['Backup']
                backup_cell.alignment = Alignment(horizontal='center')
            
            # Add borders around the entire 2x8 day block
            for row_offset in range(8):
                for col_offset in range(2):
                    current_cell = ws.cell(row=base_row + row_offset, column=col_idx + col_offset)
                    
                    # Determine which borders this cell should have
                    has_left = (col_offset == 0)
                    has_right = (col_offset == 1)
                    has_top = (row_offset == 0)
                    has_bottom = (row_offset == 7)
                    
                    # Create border object with all required sides
                    current_cell.border = Border(
                        left=thin_border if has_left else None,
                        right=thin_border if has_right else None,
                        top=thin_border if has_top else None,
                        bottom=thin_border if has_bottom else None
                    )
            
            # Remove any week separator borders that might interfere with day blocks
            if week_num > 0:
                for col_offset in range(2):
                    separator_cell = ws.cell(row=base_row - 1, column=col_idx + col_offset)
                    separator_cell.border = Border()
    
    return ws

# Only keep the function definitions, remove the file saving code
if __name__ == '__main__':
    pass 