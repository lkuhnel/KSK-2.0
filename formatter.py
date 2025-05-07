# formatter.py
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO

def format_schedule_excel(blank_excel_file, schedule_df):
    # Load template from the uploaded file
    template_wb = load_workbook(filename=BytesIO(blank_excel_file.getvalue()))

    # Dynamically create sheet dictionary for all months July 2025 to June 2026
    months = [
        ("July", 2025), ("August", 2025), ("September", 2025), ("October", 2025),
        ("November", 2025), ("December", 2025), ("January", 2026), ("February", 2026),
        ("March", 2026), ("April", 2026), ("May", 2026), ("June", 2026)
    ]
    
    sheets = {}
    for month, year in months:
        sheet_name = f"{month} {year}"
        if sheet_name in template_wb.sheetnames:
            sheets[sheet_name] = template_wb[sheet_name]

    # Calculate start weekday for each month
    month_start_weekday = {}
    for month, year in months:
        dt = datetime(year, datetime.strptime(month, "%B").month, 1)
        month_start_weekday[f"{month} {year}"] = (dt.weekday() + 1) % 7

    # Helper to map date to correct cell
    def get_cell_for_date_dynamic(date_obj):
        month = date_obj.strftime("%B")
        year = date_obj.year
        sheet_key = f"{month} {year}"
        if sheet_key not in sheets:
            return None, None, None
        start_weekday = month_start_weekday[sheet_key]
        day = date_obj.day
        total_days = day + start_weekday - 1
        week_number = total_days // 7
        day_of_week_index = total_days % 7
        base_row = 3 + (week_number * 8)
        col = day_of_week_index * 2 + 1
        return base_row, col, sheet_key

    # Clear only call (row+3) and backup (row+7) for all months
    for sheet in sheets.values():
        for week_start in range(3, 50, 8):
            for dow_col in range(1, 14, 2):
                sheet.cell(row=week_start + 3, column=dow_col).value = None  # Clear call
                sheet.cell(row=week_start + 7, column=dow_col).value = None  # Clear backup

    # Insert call and backup data
    for idx, row in schedule_df.iterrows():
        current_date = datetime.strptime(row['Date'], "%Y-%m-%d")
        call = row['Call']
        backup = row['Backup']
        base_row, col, sheet_key = get_cell_for_date_dynamic(current_date)
        if sheet_key and base_row and col:
            sheet = sheets[sheet_key]
            sheet.cell(row=base_row + 3, column=col, value=call)
            sheet.cell(row=base_row + 7, column=col, value=backup)

    # Save to BytesIO object for Streamlit
    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output
